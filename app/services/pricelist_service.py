import io
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from app.core.db_utils import get_planning_session, get_mysql_session
from app.models import planning_models
from app.models.estate_models import EstateSell, EstateHouse
from app.models.planning_models import map_russian_to_mysql_key, PropertyType
from app.services import currency_service

RESERVATION_FEE = 3_000_000
REMAINDER_STATUSES = ["Маркетинговый резерв", "Подбор"]


def calculate_new_prices(complex_name, property_type_ru, percent_change, excluded_ids=None):
    """
    Рассчитывает новые цены с учетом списка исключенных ID.
    Для объектов в списке excluded_ids повышение цены не применяется (изменение = 0%).
    """
    if excluded_ids is None:
        excluded_ids = []

    mysql_session = get_mysql_session()
    planning_session = get_planning_session()
    try:
        usd_rate = currency_service.get_current_effective_rate() or 12800.0
        prop_type_enum = PropertyType(property_type_ru)
        mysql_key = map_russian_to_mysql_key(property_type_ru)

        active_version = planning_session.query(planning_models.DiscountVersion).filter_by(is_active=True).first()
        if not active_version:
            return None, "Нет активной версии скидок"

        discount = planning_session.query(planning_models.Discount).filter_by(
            version_id=active_version.id,
            complex_name=complex_name,
            property_type=prop_type_enum,
            payment_method=planning_models.PaymentMethod.FULL_PAYMENT
        ).first()

        total_discount_rate = (discount.mpp or 0) + (discount.rop or 0) + (discount.kd or 0) if discount else 0
        discount_multiplier = 1 - total_discount_rate

        houses = mysql_session.query(EstateHouse).filter_by(complex_name=complex_name).all()
        house_map = {h.id: h.name for h in houses}
        house_ids = list(house_map.keys())

        all_objects = mysql_session.query(EstateSell).filter(
            EstateSell.house_id.in_(house_ids),
            EstateSell.estate_sell_category == mysql_key
        ).all()

        excel_results = []
        stats = {
            'complex_name': complex_name,
            'prop_type': property_type_ru,
            'total_units_project': len(all_objects),
            'remainder_units': 0,
            'remainder_sqm': 0.0,
            'discount_pct': round(total_discount_rate * 100, 2),
            'percent_change': round(percent_change * 100, 2),
            'floor_prices_before': [], 'floor_prices_after': [],
            'floor_totals_before': [], 'floor_totals_after': [],
            'final_totals_before': 0.0, 'final_totals_after': 0.0,
            'usd_rate': usd_rate,
            'by_house_rooms': {},
            'by_floor': {},
            'by_rooms_comparison': {}
        }

        for obj in all_objects:
            if not obj.estate_price or not obj.estate_area or obj.estate_area <= 0:
                continue

            # Определение процента изменения для конкретного объекта:
            # Если ID в списке исключений, используем 0, иначе стандартный процент.
            current_unit_percent = 0 if obj.id in excluded_ids else percent_change

            c_floor_total_uzs = (obj.estate_price - RESERVATION_FEE) * discount_multiplier
            c_floor_sqm_usd = (c_floor_total_uzs / obj.estate_area) / usd_rate

            # Применение индивидуального процента к цене за кв.м.
            n_floor_sqm_usd = c_floor_sqm_usd * (1 + current_unit_percent)

            n_floor_total_uzs = (n_floor_sqm_usd * usd_rate) * obj.estate_area
            n_estate_price = (n_floor_total_uzs / discount_multiplier) + RESERVATION_FEE

            excel_results.append({'Id обьекта': obj.id, 'Новая стоимость': round(n_estate_price, -3)})

            if obj.estate_sell_status_name in REMAINDER_STATUSES:
                stats['remainder_units'] += 1
                stats['remainder_sqm'] += obj.estate_area
                stats['floor_prices_before'].append(c_floor_sqm_usd)
                stats['floor_prices_after'].append(n_floor_sqm_usd)
                stats['floor_totals_before'].append(c_floor_total_uzs / usd_rate)
                stats['floor_totals_after'].append(n_floor_total_uzs / usd_rate)
                stats['final_totals_before'] += obj.estate_price / usd_rate
                stats['final_totals_after'] += n_estate_price / usd_rate

                h_name = house_map.get(obj.house_id, "Неизвестно")
                hr_key = (h_name, obj.estate_rooms or 0)
                if hr_key not in stats['by_house_rooms']:
                    stats['by_house_rooms'][hr_key] = []
                stats['by_house_rooms'][hr_key].append(n_floor_sqm_usd)

                fl_key = obj.estate_floor or 0
                if fl_key not in stats['by_floor']:
                    stats['by_floor'][fl_key] = []
                stats['by_floor'][fl_key].append(n_floor_sqm_usd)

                rooms = obj.estate_rooms or 0
                if rooms not in stats['by_rooms_comparison']:
                    stats['by_rooms_comparison'][rooms] = {'before': [], 'after': []}
                stats['by_rooms_comparison'][rooms]['before'].append(c_floor_sqm_usd)
                stats['by_rooms_comparison'][rooms]['after'].append(n_floor_sqm_usd)

        return excel_results, stats
    finally:
        mysql_session.close()
        planning_session.close()


def _set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def generate_pricelist_excel(results, stats):
    output = io.BytesIO()
    wb = Workbook()

    # Лист 1: Технический реестр
    ws1 = wb.active
    ws1.title = "PriceList"
    ws1.append(["Id обьекта", "Новая стоимость"])
    for res in results:
        ws1.append([res['Id обьекта'], res['Новая стоимость']])

    # Лист 2: На подпись
    ws = wb.create_sheet("На_подпись")

    f_header = Font(bold=True, name='Arial', size=11)
    f_bold = Font(bold=True, name='Arial', size=10)
    a_center = Alignment(horizontal='center', vertical='center')
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    for col in ['F', 'G', 'H', 'I', 'J', 'L', 'M', 'N', 'O']:
        ws.column_dimensions[col].width = 15

    # 1. Метаданные
    metadata = [
        ("Проект", stats['complex_name']),
        ("Тип недвижимости", stats['prop_type']),
        ("Всего квартир в проекте, шт.", stats['total_units_project']),
        ("Всего товарного запаса, шт", stats['remainder_units']),
        ("Всего товарного запаса, кв.м", round(stats['remainder_sqm'], 2)),
        ("Сумма регламентных скидок, %", f"{stats['discount_pct']}%"),
        ("Средняя площадь товарного запаса, кв.м",
         round(stats['remainder_sqm'] / stats['remainder_units'], 2) if stats['remainder_units'] > 0 else 0),
        ("Процент повышения цены дна, %", f"{stats['percent_change']}%")
    ]

    for i, (label, value) in enumerate(metadata, 2):
        ws.cell(row=i, column=2, value=label).font = f_bold
        ws.cell(row=i, column=2).fill = fill_gray
        ws.cell(row=i, column=3, value=value)
        _set_border(ws, f'B{i}:C{i}')

    # 2. Общее изменение цены
    ws.merge_cells('B11:D11')
    hdr = ws['B11']
    hdr.value = "Изменение цены"
    hdr.font = f_header
    hdr.alignment = a_center
    hdr.fill = fill_yellow
    _set_border(ws, 'B11:D11')

    ws['C13'], ws['D13'] = "Было", "Стало"
    for col in [3, 4]:
        cell = ws.cell(row=13, column=col)
        cell.font = f_bold
        cell.alignment = a_center
        cell.fill = fill_gray
    _set_border(ws, 'C13:D13')

    # Цена дна
    for idx, label in enumerate(["Минимальная цена дна, $", "Средняя цена дна, $", "Максимальная цена дна, $"]):
        row_idx = 14 + idx
        ws.cell(row=row_idx, column=2, value=label).font = f_bold
        before, after = stats['floor_prices_before'], stats['floor_prices_after']
        if before:
            vals = [min(before), sum(before) / len(before), max(before)]
            n_vals = [min(after), sum(after) / len(after), max(after)]
            ws.cell(row=row_idx, column=3, value=round(vals[idx], 2))
            ws.cell(row=row_idx, column=4, value=round(n_vals[idx], 2))
        _set_border(ws, f'B{row_idx}:D{row_idx}')

    # Стоимость дна
    for idx, label in enumerate(
            ["Минимальная стоимость дна, $", "Средняя стоимость дна, $", "Максимальная стоимость дна, $"]):
        row_idx = 18 + idx
        ws.cell(row=row_idx, column=2, value=label).font = f_bold
        before, after = stats['floor_totals_before'], stats['floor_totals_after']
        if before:
            vals = [min(before), sum(before) / len(before), max(before)]
            n_vals = [min(after), sum(after) / len(after), max(after)]
            ws.cell(row=row_idx, column=3, value=round(vals[idx], 0))
            ws.cell(row=row_idx, column=4, value=round(n_vals[idx], 0))
        _set_border(ws, f'B{row_idx}:D{row_idx}')

    ws.cell(row=22, column=2, value="Общая стоимость остатков, $").font = f_header
    ws.cell(row=22, column=3, value=round(stats['final_totals_before'], 0))
    ws.cell(row=22, column=4, value=round(stats['final_totals_after'], 0))
    _set_border(ws, 'B22:D22')

    # --- НОВАЯ ТАБЛИЦА: Типология (Было/Стало) ---
    start_row_typ = 25
    ws.merge_cells(f'B{start_row_typ}:H{start_row_typ}')
    typ_hdr = ws.cell(row=start_row_typ, column=2, value="Сравнение по типологии (цена дна за кв.м, $)")
    typ_hdr.font = f_header
    typ_hdr.alignment = a_center
    typ_hdr.fill = fill_yellow
    _set_border(ws, f'B{start_row_typ}:H{start_row_typ}')

    h_row = start_row_typ + 1
    sub_headers = ["Комн.", "Мин (Было)", "Мин (Стало)", "Сред (Было)", "Сред (Стало)", "Макс (Было)", "Макс (Стало)"]
    for i, text in enumerate(sub_headers):
        cell = ws.cell(row=h_row, column=2 + i, value=text)
        cell.font = f_bold
        cell.fill = fill_gray
        cell.alignment = a_center
    _set_border(ws, f'B{h_row}:H{h_row}')

    curr_row = h_row + 1
    for rooms in sorted(stats['by_rooms_comparison'].keys()):
        data = stats['by_rooms_comparison'][rooms]
        b_vals, a_vals = data['before'], data['after']

        ws.cell(row=curr_row, column=2, value=f"{rooms}-комн")
        ws.cell(row=curr_row, column=3, value=round(min(b_vals), 2))
        ws.cell(row=curr_row, column=4, value=round(min(a_vals), 2))
        ws.cell(row=curr_row, column=5, value=round(sum(b_vals) / len(b_vals), 2))
        ws.cell(row=curr_row, column=6, value=round(sum(a_vals) / len(a_vals), 2))
        ws.cell(row=curr_row, column=7, value=round(max(b_vals), 2))
        ws.cell(row=curr_row, column=8, value=round(max(a_vals), 2))

        _set_border(ws, f'B{curr_row}:H{curr_row}')
        curr_row += 1

    # --- Таблица справа 1: По домам и комнатности ---
    s_col, row = 6, 2
    headers = ["Дом", "Комн.", "Мин. цена $", "Сред. цена $", "Макс. цена $"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=s_col + i, value=h)
        cell.font = f_bold
        cell.fill = fill_gray
        cell.alignment = a_center
    _set_border(ws, f'F{row}:J{row}')

    row += 1
    sorted_hr = sorted(stats['by_house_rooms'].items(), key=lambda x: (x[0][0], x[0][1]))
    for (h_name, rooms), prices in sorted_hr:
        ws.cell(row=row, column=s_col, value=h_name)
        ws.cell(row=row, column=s_col + 1, value=rooms)
        ws.cell(row=row, column=s_col + 2, value=round(min(prices), 2))
        ws.cell(row=row, column=s_col + 3, value=round(sum(prices) / len(prices), 2))
        ws.cell(row=row, column=s_col + 4, value=round(max(prices), 2))
        _set_border(ws, f'F{row}:J{row}')
        row += 1

    # --- Таблица справа 2: По этажам ---
    s_col_fl, row_fl = 12, 2
    fl_headers = ["Этаж", "Мин. цена $", "Сред. цена $", "Макс. цена $"]
    for i, h in enumerate(fl_headers):
        cell = ws.cell(row=row_fl, column=s_col_fl + i, value=h)
        cell.font = f_bold
        cell.fill = fill_gray
        cell.alignment = a_center
    _set_border(ws, f'L{row_fl}:O{row_fl}')

    row_fl += 1
    sorted_fl = sorted(stats['by_floor'].items(), key=lambda x: x[0])
    for floor, prices in sorted_fl:
        ws.cell(row=row_fl, column=s_col_fl, value=floor)
        ws.cell(row=row_fl, column=s_col_fl + 1, value=round(min(prices), 2))
        ws.cell(row=row_fl, column=s_col_fl + 2, value=round(sum(prices) / len(prices), 2))
        ws.cell(row=row_fl, column=s_col_fl + 3, value=round(max(prices), 2))
        _set_border(ws, f'L{row_fl}:O{row_fl}')
        row_fl += 1

    wb.save(output)
    output.seek(0)
    return output